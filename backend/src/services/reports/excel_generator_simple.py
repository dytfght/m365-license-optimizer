"""
Simple Excel Report Generator - Creates Excel reports without pandas dependency
"""
import io
from datetime import datetime
from typing import Any, Dict

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

logger = structlog.get_logger(__name__)


class ExcelGenerator:
    """Generate detailed Excel reports with Microsoft design"""

    # Microsoft colors
    COLORS = {
        "primary": "0078D4",  # Microsoft blue
        "background": "F3F2F1",  # Light gray
        "text": "323130",  # Dark gray
        "text_secondary": "605E5C",  # Medium gray
        "success": "107C10",  # Green
        "warning": "FF8C00",  # Orange
        "error": "D13438",  # Red
        "white": "FFFFFF",
        "black": "000000",
    }

    def __init__(self):
        self.setup_styles()

    def setup_styles(self):
        """Setup Excel styles for Microsoft design"""

        # Just define colors for now - keep it simple
        self.green_fill = PatternFill(
            start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"
        )
        self.red_fill = PatternFill(
            start_color="FFE8E8", end_color="FFE8E8", fill_type="solid"
        )
        self.yellow_fill = PatternFill(
            start_color="FFF8E8", end_color="FFF8E8", fill_type="solid"
        )

    async def generate_detailed_excel(self, data: Dict[str, Any]) -> bytes:
        """Generate detailed Excel report with 3 sheets"""

        logger.info("generating_excel_report", analysis_id=data.get("analysis_id"))

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create 3 sheets
        wb.create_sheet("Synthèse", 0)
        wb.create_sheet("Recommandations détaillées", 1)
        wb.create_sheet("Données brutes", 2)

        # Generate each sheet
        self._create_summary_sheet(wb, data)
        self._create_detailed_sheet(wb, data)
        self._create_raw_data_sheet(wb, data)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info("excel_generated_successfully", size_bytes=len(buffer.getvalue()))

        return buffer.getvalue()

    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create summary sheet with KPIs and charts"""

        ws = wb["Synthèse"]

        # Title
        ws["A1"] = "SYNTHÈSE - ANALYSE D'OPTIMISATION MICROSOFT 365"
        ws["A1"].font = Font(
            name="Segoe UI", size=16, bold=True, color=self.COLORS["primary"]
        )
        ws.merge_cells("A1:H1")

        # Date
        ws["A3"] = f"Date de génération: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["A3"].font = Font(
            name="Segoe UI", size=10, color=self.COLORS["text_secondary"]
        )

        # KPI Section (A5:B10)
        ws["A5"] = "INDICATEURS CLÉS"
        ws["A5"].font = Font(
            name="Segoe UI", size=12, bold=True, color=self.COLORS["primary"]
        )

        kpis = data.get("kpis", {})
        kpi_data = [
            ["Coût actuel mensuel", kpis.get("current_monthly_cost", 0)],
            ["Coût cible mensuel", kpis.get("target_monthly_cost", 0)],
            ["Économie mensuelle", kpis.get("monthly_savings", 0)],
            ["Économie annuelle", kpis.get("annual_savings", 0)],
            ["Taux d'économie", kpis.get("savings_percentage", 0)],
        ]

        for i, (label, value) in enumerate(kpi_data, start=6):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"B{i}"].number_format = "#,##0.00 €"
            ws[f"B{i}"].font = Font(name="Segoe UI", size=11)
            ws[f"B{i}"].alignment = Alignment(horizontal="right")

            # Color code savings
            if "Économie" in label and value > 0:
                ws[f"B{i}"].fill = self.green_fill

        # Chart section (D5:H15)
        ws["D5"] = "RÉPARTITION DES LICENCES"
        ws["D5"].font = Font(
            name="Segoe UI", size=12, bold=True, color=self.COLORS["primary"]
        )

        # Prepare data for chart
        license_data = data.get("license_distribution", [])

        if license_data:
            # Write data for chart
            ws["D7"] = "Licence"
            ws["E7"] = "Utilisateurs"
            ws["F7"] = "Pourcentage"

            for i, item in enumerate(license_data):
                ws[f"D{i+8}"] = item.get("license_name", f"License {i}")
                ws[f"E{i+8}"] = item.get("user_count", 0)
                ws[f"F{i+8}"] = item.get("percentage", 0) / 100
                ws[f"F{i+8}"].number_format = "0.00%"

        # No data validation for now - keep it simple

    def _create_detailed_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create detailed recommendations sheet"""

        ws = wb["Recommandations détaillées"]

        # Title
        ws["A1"] = "RECOMMANDATIONS DÉTAILLÉES PAR UTILISATEUR"
        ws["A1"].font = Font(
            name="Segoe UI", size=14, bold=True, color=self.COLORS["primary"]
        )
        ws.merge_cells("A1:R1")

        # Headers
        headers = [
            "UserPrincipalName",
            "DisplayName",
            "Department",
            "JobTitle",
            "CurrentLicense",
            "CurrentMonthlyCost",
            "RecommendedLicense",
            "RecommendedMonthlyCost",
            "MonthlySavings",
            "AnnualSavings",
            "ExchangeUsageLevel",
            "OneDriveUsageGB",
            "TeamsActivity",
            "OfficeDesktopRequired",
            "LastActivityDate",
            "InactivityDays",
            "RecommendationStatus",
            "RecommendationReason",
        ]

        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=header)
            cell.font = Font(
                name="Segoe UI", size=11, bold=True, color=self.COLORS["white"]
            )
            cell.fill = PatternFill(
                start_color=self.COLORS["primary"],
                end_color=self.COLORS["primary"],
                fill_type="solid",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Get recommendations data
        recommendations = data.get("recommendations", [])

        # Write data
        for i, rec in enumerate(recommendations, start=4):
            ws.cell(row=i, column=1, value=rec.get("user_principal_name", ""))
            ws.cell(row=i, column=2, value=rec.get("display_name", ""))
            ws.cell(row=i, column=3, value=rec.get("department", ""))
            ws.cell(row=i, column=4, value=rec.get("job_title", ""))
            ws.cell(row=i, column=5, value=rec.get("current_license", ""))
            ws.cell(row=i, column=6, value=rec.get("current_monthly_cost", 0))
            ws.cell(row=i, column=6).number_format = "#,##0.00 €"
            ws.cell(row=i, column=7, value=rec.get("recommended_license", ""))
            ws.cell(row=i, column=8, value=rec.get("recommended_monthly_cost", 0))
            ws.cell(row=i, column=8).number_format = "#,##0.00 €"

            # Savings with conditional formatting
            monthly_savings = rec.get("monthly_savings", 0)
            ws.cell(row=i, column=9, value=monthly_savings)
            ws.cell(row=i, column=9).number_format = "#,##0.00 €"
            if monthly_savings > 0:
                ws.cell(row=i, column=9).fill = self.green_fill

            annual_savings = rec.get("annual_savings", 0)
            ws.cell(row=i, column=10, value=annual_savings)
            ws.cell(row=i, column=10).number_format = "#,##0.00 €"
            if annual_savings > 0:
                ws.cell(row=i, column=10).fill = self.green_fill

            # Usage levels
            ws.cell(row=i, column=11, value=rec.get("exchange_usage_level", "Faible"))
            ws.cell(row=i, column=12, value=rec.get("onedrive_usage_gb", 0))
            ws.cell(row=i, column=13, value=rec.get("teams_activity", "Faible"))
            ws.cell(
                row=i,
                column=14,
                value="Oui" if rec.get("office_desktop_required", False) else "Non",
            )

            # Activity dates
            last_activity = rec.get("last_activity_date")
            if last_activity:
                ws.cell(row=i, column=15, value=last_activity)
                ws.cell(row=i, column=15).number_format = "DD/MM/YYYY"

            inactivity_days = rec.get("inactivity_days", 0)
            ws.cell(row=i, column=16, value=inactivity_days)
            if inactivity_days > 90:
                ws.cell(row=i, column=16).fill = self.red_fill

            # Recommendation status
            ws.cell(row=i, column=17, value=rec.get("recommendation_status", "Proposé"))
            ws.cell(row=i, column=18, value=rec.get("recommendation_reason", ""))

        # Auto-filter (simplified)
        # ws.auto_filter.ref = f"A3:R{3+len(recommendations)}"  # Commented out for compatibility

        # Freeze first row
        ws.freeze_panes = "A4"

        # Freeze first row
        ws.freeze_panes = "A4"

        # Column widths set manually for simplicity
        ws.column_dimensions["A"].width = 30  # UserPrincipalName
        ws.column_dimensions["B"].width = 25  # DisplayName
        ws.column_dimensions["C"].width = 20  # Department
        ws.column_dimensions["D"].width = 25  # JobTitle
        ws.column_dimensions["E"].width = 30  # CurrentLicense
        ws.column_dimensions["F"].width = 15  # CurrentMonthlyCost
        ws.column_dimensions["G"].width = 30  # RecommendedLicense
        ws.column_dimensions["H"].width = 18  # RecommendedMonthlyCost
        ws.column_dimensions["I"].width = 15  # MonthlySavings
        ws.column_dimensions["J"].width = 15  # AnnualSavings
        ws.column_dimensions["K"].width = 18  # ExchangeUsageLevel
        ws.column_dimensions["L"].width = 18  # OneDriveUsageGB
        ws.column_dimensions["M"].width = 15  # TeamsActivity
        ws.column_dimensions["N"].width = 20  # OfficeDesktopRequired
        ws.column_dimensions["O"].width = 18  # LastActivityDate
        ws.column_dimensions["P"].width = 15  # InactivityDays
        ws.column_dimensions["Q"].width = 22  # RecommendationStatus
        ws.column_dimensions["R"].width = 40  # RecommendationReason

    def _add_data_validations(self, ws):
        """Add data validation for dropdown lists"""

        # Usage levels
        usage_validation = DataValidation(
            type="list", formula1='"Faible,Moyen,Élevé,Aucun"', allow_blank=True
        )
        usage_validation.error = "Veuillez sélectionner une valeur valide"
        usage_validation.errorTitle = "Validation erreur"
        ws.add_data_validation(usage_validation)
        usage_validation.add(f"K4:K{ws.max_row}")
        usage_validation.add(f"M4:M{ws.max_row}")

        # Boolean values
        bool_validation = DataValidation(
            type="list", formula1='"Oui,Non"', allow_blank=True
        )
        ws.add_data_validation(bool_validation)
        bool_validation.add(f"N4:N{ws.max_row}")

        # Recommendation status
        status_validation = DataValidation(
            type="list", formula1='"Proposé,Validé,Rejeté,Sensible"', allow_blank=True
        )
        ws.add_data_validation(status_validation)
        status_validation.add(f"Q4:Q{ws.max_row}")

    def _create_raw_data_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create raw data sheet for advanced analysis"""

        ws = wb["Données brutes"]

        # Title
        ws["A1"] = "DONNÉES BRUTES D'ANALYSE"
        ws["A1"].font = Font(
            name="Segoe UI", size=12, bold=True, color=self.COLORS["primary"]
        )

        # For testing, just add some sample data
        ws["A3"] = "Type de données"
        ws["B3"] = "Valeur"
        ws["C3"] = "Date"

        sample_data = [
            [
                "Total utilisateurs",
                data.get("kpis", {}).get("total_users", 0),
                datetime.now(),
            ],
            [
                "Économies mensuelles",
                data.get("kpis", {}).get("monthly_savings", 0),
                datetime.now(),
            ],
            [
                "Économies annuelles",
                data.get("kpis", {}).get("annual_savings", 0),
                datetime.now(),
            ],
        ]

        for i, (label, value, date) in enumerate(sample_data, start=4):
            ws.cell(row=i, column=1, value=label)
            ws.cell(row=i, column=2, value=value)
            ws.cell(row=i, column=3, value=date)
            ws.cell(row=i, column=3).number_format = "DD/MM/YYYY"

        # Column widths for summary sheet
        ws.column_dimensions["A"].width = 25  # KPI labels
        ws.column_dimensions["B"].width = 15  # KPI values
