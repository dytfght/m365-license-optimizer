"""
Excel Report Generator - Creates detailed Excel reports with formatting
"""
import io
from datetime import datetime
from typing import Any, Dict

import pandas as pd
import structlog
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

logger = structlog.get_logger(__name__)


class ExcelGenerator:
    """Generate detailed Excel reports with Microsoft design"""

    # Microsoft colors
    COLORS = {
        "primary": "0078D4",  # Microsoft blue
        "background": "F3F2F1",  # Light gray
        "text": "323130",  # Dark gray
        "text_secondary": "605E5C",  # Medium gray
        "success": "107C10",  # Green
        "warning": "FF8C00",  # Orange
        "error": "D13438",  # Red
        "white": "FFFFFF",
        "black": "000000",
    }

    def __init__(self):
        self.setup_styles()

    def setup_styles(self):
        """Setup Excel styles for Microsoft design"""

        # Header style
        self.header_style = NamedStyle(name="header")
        self.header_style.font = Font(
            name="Segoe UI", size=12, bold=True, color=self.COLORS["white"]
        )
        self.header_style.fill = PatternFill(
            start_color=self.COLORS["primary"],
            end_color=self.COLORS["primary"],
            fill_type="solid",
        )
        self.header_style.alignment = Alignment(horizontal="center", vertical="center")
        self.header_style.border = Border(
            left=Side(style="thin", color=self.COLORS["text_secondary"]),
            right=Side(style="thin", color=self.COLORS["text_secondary"]),
            top=Side(style="thin", color=self.COLORS["text_secondary"]),
            bottom=Side(style="thin", color=self.COLORS["text_secondary"]),
        )

        # KPI style
        self.kpi_value_style = NamedStyle(name="kpi_value")
        self.kpi_value_style.font = Font(
            name="Segoe UI", size=14, bold=True, color=self.COLORS["primary"]
        )
        self.kpi_value_style.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        self.kpi_value_style.number_format = "#,##0.00 €"

        # Currency style
        self.currency_style = NamedStyle(name="currency")
        self.currency_style.number_format = "#,##0.00 €"
        self.currency_style.font = Font(name="Segoe UI", size=11)
        self.currency_style.alignment = Alignment(horizontal="right", vertical="center")

        # Date style
        self.date_style = NamedStyle(name="date")
        self.date_style.number_format = "DD/MM/YYYY"
        self.date_style.font = Font(name="Segoe UI", size=11)
        self.date_style.alignment = Alignment(horizontal="center", vertical="center")

        # Percentage style
        self.percent_style = NamedStyle(name="percentage")
        self.percent_style.number_format = "0.00%"
        self.percent_style.font = Font(name="Segoe UI", size=11)
        self.percent_style.alignment = Alignment(horizontal="center", vertical="center")

        # Conditional formatting colors
        self.green_fill = PatternFill(
            start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"
        )
        self.red_fill = PatternFill(
            start_color="FFE8E8", end_color="FFE8E8", fill_type="solid"
        )
        self.yellow_fill = PatternFill(
            start_color="FFF8E8", end_color="FFF8E8", fill_type="solid"
        )

    async def generate_detailed_excel(self, data: Dict[str, Any]) -> bytes:
        """Generate detailed Excel report with 3 sheets"""

        logger.info("generating_excel_report", analysis_id=data.get("analysis_id"))

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create 3 sheets
        wb.create_sheet("Synthèse", 0)
        wb.create_sheet("Recommandations détaillées", 1)
        wb.create_sheet("Données brutes", 2)

        # Generate each sheet
        self._create_summary_sheet(wb, data)
        self._create_detailed_sheet(wb, data)
        self._create_raw_data_sheet(wb, data)

        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info("excel_generated_successfully", size_bytes=len(buffer.getvalue()))

        return buffer.getvalue()

    def _create_summary_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create summary sheet with KPIs and charts"""

        ws = wb["Synthèse"]

        # Add custom styles to workbook
        for style in [self.header_style, self.kpi_value_style, self.currency_style]:
            if style.name not in wb.named_styles:
                wb.add_named_style(style)

        # Title
        ws["A1"] = "SYNTHÈSE - ANALYSE D'OPTIMISATION MICROSOFT 365"
        ws["A1"].font = Font(
            name="Segoe UI", size=16, bold=True, color=self.COLORS["primary"]
        )
        ws.merge_cells("A1:H1")

        # Date
        ws["A3"] = f"Date de génération: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        ws["A3"].font = Font(
            name="Segoe UI", size=10, color=self.COLORS["text_secondary"]
        )

        # KPI Section (A5:B10)
        ws["A5"] = "INDICATEURS CLÉS"
        ws["A5"].font = Font(
            name="Segoe UI", size=12, bold=True, color=self.COLORS["primary"]
        )

        kpis = data.get("kpis", {})
        kpi_data = [
            ["Coût actuel mensuel", kpis.get("current_monthly_cost", 0)],
            ["Coût cible mensuel", kpis.get("target_monthly_cost", 0)],
            ["Économie mensuelle", kpis.get("monthly_savings", 0)],
            ["Économie annuelle", kpis.get("annual_savings", 0)],
            ["Taux d'économie", kpis.get("savings_percentage", 0)],
        ]

        for i, (label, value) in enumerate(kpi_data, start=6):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
            ws[f"B{i}"].style = "currency"

            # Color code savings
            if "Économie" in label and value > 0:
                ws[f"B{i}"].fill = self.green_fill

        # Chart section (D5:H15)
        ws["D5"] = "RÉPARTITION DES LICENCES"
        ws["D5"].font = Font(
            name="Segoe UI", size=12, bold=True, color=self.COLORS["primary"]
        )

        # Prepare data for chart
        license_data = data.get("license_distribution", [])
        chart_df = pd.DataFrame(license_data)

        if not chart_df.empty:
            # Write data for chart
            ws["D7"] = "Licence"
            ws["E7"] = "Utilisateurs"
            ws["F7"] = "Pourcentage"

            for i, row in chart_df.iterrows():
                ws[f"D{i+8}"] = row.get("license_name", f"License {i}")
                ws[f"E{i+8}"] = row.get("user_count", 0)
                ws[f"F{i+8}"] = row.get("percentage", 0) / 100
                ws[f"F{i+8}"].style = "percentage"

            # Create bar chart
            chart = BarChart()
            chart.title = "Répartition des licences par type"
            chart.style = 2
            chart.y_axis.title = "Nombre d'utilisateurs"
            chart.x_axis.title = "Type de licence"

            data_ref = Reference(
                ws, min_col=5, min_row=8, max_row=8 + len(chart_df), max_col=5
            )
            cats_ref = Reference(ws, min_col=4, min_row=8, max_row=8 + len(chart_df))

            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(cats_ref)

            ws.add_chart(chart, "D10")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_detailed_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create detailed recommendations sheet"""

        ws = wb["Recommandations détaillées"]

        # Title
        ws["A1"] = "RECOMMANDATIONS DÉTAILLÉES PAR UTILISATEUR"
        ws["A1"].font = Font(
            name="Segoe UI", size=14, bold=True, color=self.COLORS["primary"]
        )
        ws.merge_cells("A1:R1")

        # Headers
        headers = [
            "UserPrincipalName",
            "DisplayName",
            "Department",
            "JobTitle",
            "CurrentLicense",
            "CurrentMonthlyCost",
            "RecommendedLicense",
            "RecommendedMonthlyCost",
            "MonthlySavings",
            "AnnualSavings",
            "ExchangeUsageLevel",
            "OneDriveUsageGB",
            "TeamsActivity",
            "OfficeDesktopRequired",
            "LastActivityDate",
            "InactivityDays",
            "RecommendationStatus",
            "RecommendationReason",
        ]

        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=i, value=header)
            cell.style = self.header_style

        # Get recommendations data
        recommendations = data.get("recommendations", [])

        # Write data
        for i, rec in enumerate(recommendations, start=4):
            ws.cell(row=i, column=1, value=rec.get("user_principal_name", ""))
            ws.cell(row=i, column=2, value=rec.get("display_name", ""))
            ws.cell(row=i, column=3, value=rec.get("department", ""))
            ws.cell(row=i, column=4, value=rec.get("job_title", ""))
            ws.cell(row=i, column=5, value=rec.get("current_license", ""))
            ws.cell(row=i, column=6, value=rec.get("current_monthly_cost", 0))
            ws.cell(row=i, column=6).style = "currency"
            ws.cell(row=i, column=7, value=rec.get("recommended_license", ""))
            ws.cell(row=i, column=8, value=rec.get("recommended_monthly_cost", 0))
            ws.cell(row=i, column=8).style = "currency"

            # Savings with conditional formatting
            monthly_savings = rec.get("monthly_savings", 0)
            ws.cell(row=i, column=9, value=monthly_savings)
            ws.cell(row=i, column=9).style = "currency"
            if monthly_savings > 0:
                ws.cell(row=i, column=9).fill = self.green_fill

            annual_savings = rec.get("annual_savings", 0)
            ws.cell(row=i, column=10, value=annual_savings)
            ws.cell(row=i, column=10).style = "currency"
            if annual_savings > 0:
                ws.cell(row=i, column=10).fill = self.green_fill

            # Usage levels
            ws.cell(row=i, column=11, value=rec.get("exchange_usage_level", "Faible"))
            ws.cell(row=i, column=12, value=rec.get("onedrive_usage_gb", 0))
            ws.cell(row=i, column=13, value=rec.get("teams_activity", "Faible"))
            ws.cell(
                row=i,
                column=14,
                value="Oui" if rec.get("office_desktop_required", False) else "Non",
            )

            # Activity dates
            last_activity = rec.get("last_activity_date")
            if last_activity:
                ws.cell(row=i, column=15, value=last_activity)
                ws.cell(row=i, column=15).style = self.date_style

            inactivity_days = rec.get("inactivity_days", 0)
            ws.cell(row=i, column=16, value=inactivity_days)
            if inactivity_days > 90:
                ws.cell(row=i, column=16).fill = self.red_fill

            # Recommendation status
            ws.cell(row=i, column=17, value=rec.get("recommendation_status", "Proposé"))
            ws.cell(row=i, column=18, value=rec.get("recommendation_reason", ""))

        # Add data validation for dropdown lists
        self._add_data_validations(ws)

        # Auto-filter
        ws.auto_filter.ref = f"A3:R{3+len(recommendations)}"

        # Freeze first row
        ws.freeze_panes = "A4"

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_data_validations(self, ws):
        """Add data validation for dropdown lists"""

        # Usage levels
        usage_validation = DataValidation(
            type="list", formula1='"Faible,Moyen,Élevé,Aucun"', allow_blank=True
        )
        usage_validation.error = "Veuillez sélectionner une valeur valide"
        usage_validation.errorTitle = "Validation erreur"
        ws.add_data_validation(usage_validation)
        usage_validation.add(f"K4:K{ws.max_row}")
        usage_validation.add(f"M4:M{ws.max_row}")

        # Boolean values
        bool_validation = DataValidation(
            type="list", formula1='"Oui,Non"', allow_blank=True
        )
        ws.add_data_validation(bool_validation)
        bool_validation.add(f"N4:N{ws.max_row}")

        # Recommendation status
        status_validation = DataValidation(
            type="list", formula1='"Proposé,Validé,Rejeté,Sensible"', allow_blank=True
        )
        ws.add_data_validation(status_validation)
        status_validation.add(f"Q4:Q{ws.max_row}")

    def _create_raw_data_sheet(self, wb: Workbook, data: Dict[str, Any]):
        """Create raw data sheet for advanced analysis"""

        ws = wb["Données brutes"]

        # Title
        ws["A1"] = "DONNÉES BRUTES D'ANALYSE"
        ws["A1"].font = Font(
            name="Segoe UI", size=12, bold=True, color=self.COLORS["primary"]
        )

        # Export all raw metrics
        users_data = data.get("users_metrics", [])

        if users_data:
            # Convert to DataFrame for easier manipulation
            df = pd.DataFrame(users_data)

            # Write headers
            for i, column in enumerate(df.columns, 1):
                ws.cell(row=3, column=i, value=column)
                ws.cell(row=3, column=i).style = self.header_style

            # Write data
            for i, row in df.iterrows():
                for j, column in enumerate(df.columns, 1):
                    value = row[column]
                    ws.cell(row=i + 4, column=j, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _prepare_excel_data(self, data: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
        """Prepare data for Excel generation"""

        # Summary data
        summary_data = {
            "KPI": [
                "Coût actuel mensuel",
                "Coût cible mensuel",
                "Économie mensuelle",
                "Économie annuelle",
                "Taux d'économie",
                "Utilisateurs analysés",
            ],
            "Valeur": [
                data.get("kpis", {}).get("current_monthly_cost", 0),
                data.get("kpis", {}).get("target_monthly_cost", 0),
                data.get("kpis", {}).get("monthly_savings", 0),
                data.get("kpis", {}).get("annual_savings", 0),
                data.get("kpis", {}).get("savings_percentage", 0),
                data.get("kpis", {}).get("total_users", 0),
            ],
        }

        # Detailed recommendations
        recommendations = data.get("recommendations", [])
        detailed_data = []

        for rec in recommendations:
            detailed_data.append(
                {
                    "UserPrincipalName": rec.get("user_principal_name", ""),
                    "DisplayName": rec.get("display_name", ""),
                    "Department": rec.get("department", ""),
                    "JobTitle": rec.get("job_title", ""),
                    "CurrentLicense": rec.get("current_license", ""),
                    "CurrentMonthlyCost": rec.get("current_monthly_cost", 0),
                    "RecommendedLicense": rec.get("recommended_license", ""),
                    "RecommendedMonthlyCost": rec.get("recommended_monthly_cost", 0),
                    "MonthlySavings": rec.get("monthly_savings", 0),
                    "AnnualSavings": rec.get("annual_savings", 0),
                    "ExchangeUsageLevel": rec.get("exchange_usage_level", "Faible"),
                    "OneDriveUsageGB": rec.get("onedrive_usage_gb", 0),
                    "TeamsActivity": rec.get("teams_activity", "Faible"),
                    "OfficeDesktopRequired": "Oui"
                    if rec.get("office_desktop_required", False)
                    else "Non",
                    "LastActivityDate": rec.get("last_activity_date"),
                    "InactivityDays": rec.get("inactivity_days", 0),
                    "RecommendationStatus": rec.get("recommendation_status", "Proposé"),
                    "RecommendationReason": rec.get("recommendation_reason", ""),
                }
            )

        return {
            "summary": pd.DataFrame(summary_data),
            "detailed": pd.DataFrame(detailed_data),
            "raw": pd.DataFrame(data.get("users_metrics", [])),
        }
